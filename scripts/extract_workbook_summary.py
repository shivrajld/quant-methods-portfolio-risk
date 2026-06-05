from pathlib import Path

import openpyxl
import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "model" / "quant_methods_portfolio_model.xlsx"


def workbook_overview(path: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    rows = []
    for ws in wb.worksheets:
        nonempty = 0
        formulas = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                nonempty += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas += 1
        rows.append(
            {
                "sheet": ws.title,
                "rows": ws.max_row,
                "columns": ws.max_column,
                "nonempty_cells": nonempty,
                "formula_cells": formulas,
            }
        )
    return pd.DataFrame(rows)


def main() -> None:
    overview = workbook_overview(WORKBOOK)
    print("Workbook:", WORKBOOK)
    print()
    print(overview.to_string(index=False))

    out = ROOT / "tables" / "workbook_overview.csv"
    overview.to_csv(out, index=False)
    print()
    print("Wrote:", out)


if __name__ == "__main__":
    main()
